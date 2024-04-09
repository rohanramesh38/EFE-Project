import ast
import boto3
import logging
import numpy as np
import os
import pandas as pd
import re
from openpyxl import load_workbook
import requests
import json
import streamlit as st
from botocore.exceptions import ClientError
from collections import defaultdict
from data_extractor import get_all_companys_data, extract_pages_with_text_pypdf, extract_pages_with_text_llama_parse
from fuzzywuzzy import fuzz
from io import BytesIO
import PyPDF2
from rich.console import Console
from rich.markdown import Markdown
import fitz
from openpyxl.utils.dataframe import dataframe_to_rows


console = Console()

HEADERS = {'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 \ (KHTML, like Gecko) Chrome/44.0.2403.157 Safari/537.36','Accept-Language': 'en-US, en;q=0.5'}



logger = logging.getLogger(__name__)


class BedrockWrapper:

    def __init__(self, bedrock_client):
        self.bedrock_client = bedrock_client

    def list_foundation_models(self):
        try:
            response = self.bedrock_client.list_foundation_models()
            models = response["modelSummaries"]
            logger.info("Got %s foundation models.", len(models))
        except ClientError:
            logger.error("Couldn't list foundation models.")
            raise
        else:
            return models

def get_runtime():

    logging.basicConfig(level=logging.INFO)
    bedrock_client = boto3.client(service_name="bedrock", region_name="us-east-1")

    wrapper = BedrockWrapper(bedrock_client)
    client = boto3.client('bedrock-runtime')

    return client

def generate_data(text_from_pages,prompt,timeline):
    data=''

    if(timeline):
        data=f'\n\nHuman: I m going to give you a table text extracted from a pdf and asking you to analyze the following information and extract from the key data features \\ntranscript.\\nHere is the transcript:\\n<transcript>\\n {text_from_pages} \\n </transcript> \\nYour task is to analyze that transcript, {prompt} {timeline} \\n\\nAssistant:' 

    else:    
        data=f'\n\nHuman: I m going to give you a table text extracted from a pdf and asking you to analyze the following information and extract from the key data features \\ntranscript.\\nHere is the transcript:\\n<transcript>\\n {text_from_pages} \\n </transcript> \\nYour task is to analyze that transcript, {prompt} \\n\\nAssistant:' 

    #print(data)
    return data

def generate_body(data):
    body = json.dumps({
    'prompt': data,
    "max_tokens_to_sample": 1000,
    "temperature": 0.1,
    "top_p": 0.1,
    })
    return body

def invoke_claude_model(bedrock_client,body,comany):
    modelId = 'anthropic.claude-v2'
    accept = 'application/json'
    contentType = 'application/json'

    response = bedrock_client.invoke_model(body=body, modelId=modelId, accept=accept, contentType=contentType)

    response_body = json.loads(response.get('body').read())

    result=response_body.get('completion')
    output=fappend_csv(result,sheet=comany)

    console.print(Markdown(result))

    return output


    


#Setting up thresholds to extract pages

gs_search_words_with_threshold=[('Assets Under Supervision',50),('Earnings Results',80),('AUS',100),('Wealth Management',60)]
schwab_search_words_with_threshold = [('Total asset management',50),('Total client assets',80),('Net New Assets',30) ]
bofa_search_words_with_threshold = [('Wealth Management',70,),('Assets  ',70),('Total assets',100) ]
mgs_search_words_with_threshold =[('Wealth Management Metrics',80),('Advisor‐led channel',60),('Financial Information',40),('Asset Management',50)]
jpm_search_words_with_threshold =[('Asset & Wealth',80),('Asset',60),('Wealth Management',60)]


#Setting up prompts

gs_prompt='''From the given information, create a markdown table with containing wealth management information about Wealth AUS,Wealth net flows, Wealth Management Fees, 
Net market appreciation / (depreciation)and Total Revenues,Total AUS,Wealth net flows,Wealth Mgmt Fees,Total Revenues , dont give an text other than the table and in the markdown table use  'B' for billion and 'M' for million for the quarter or months ending for the quarter'''

schwab_prompt='''From the given information, create a markdown table with containing Asset management information bout Total net new assets, Net growth in client assets, Total Advice Solutions,Total Client Assets,Investor Services Assets,Advisor Services Assets,Advice Solutions (fee-Based) Assets,Total NNA,Investor Services NNA,
Advisor Services NNA,Active Brokerage Accts (MMs), dont give an text other than the table and in the markdown table use  'B' for billion and 'M' for million for the quarter or months ending for the quarter'''

bofa_prompt='''From the given information, create a markdown table with containing Global Wealth & Investment Management information, 
Asset under management, Net Cliet Flow, Total Wealth Advisors.  dont give an text other than the table and in the markdown table use  'B' for billion and 'M' for million'''

mgs_prompt='''From the given information, create a markdown table with containing wealth management information, get Total client assets, Net new assets 
Advisor‐led client assets, Fee‐based client assets and flows , Self‐directed assets. dont give an text other than the table and in the markdown table use  'B' for billion and 'M' for million'''

jpm_prompt='''From the given information, create a markdown table with containing wealth management information of these metrics [Total client assets, Net new assets 
Total Assets under management and Assent under supervision,total,net revenue, net income,Asset management fees] Total AUM (incl. Institutional),Private Wealth AUM,Liquidity+Fixed Income+Equity+Multi-asset=Net Asset Flows (incl Institutional) and Market Impact, give rows only if value exists for a metric or extract metric in similar context. dont give an text other than the table and in the markdown table use  'B' for billion and 'M' for million'''



def result2():
    schwab_url='https://content.schwab.com/web/retail/public/about-schwab/schw_q4_2023_earnings_release.pdf'
    gs_2Q23_url='https://www.goldmansachs.com/media-relations/press-releases/current/pdfs/2023-q2-results.pdf'
    bofa_4Q2023_url='https://d1io3yog0oux5.cloudfront.net/_e26f734f80404bf9f15939064dbef560/bankofamerica/db/806/9995/supplemental_information/The+Supplemental+Information_4Q23_ADA.pdf'

    matching_pages,text=extract_pages_with_text_pypdf(bofa_4Q2023_url,bofa_search_words_with_threshold)
    print(matching_pages,bofa_4Q2023_url)
    #invoke(text,bofa_prompt)

    matching_pages,text=extract_pages_with_text_pypdf(gs_2Q23_url,gs_search_words_with_threshold)
    print(matching_pages,gs_2Q23_url)
    #invoke(text,gs_prompt)

    matching_pages,text=extract_pages_with_text_pypdf(schwab_url,schwab_search_words_with_threshold)
    print(matching_pages,schwab_url)
    invoke(text,gs_prompt)

def parse_rmd_table(table_string):
    # Find all rows
    rows = table_string.strip().split('\n')
    # Remove leading and trailing '|' and split by '|'
    rows = [re.split(r'\s*\|\s*', row.strip('|')) for row in rows]
    # Extract header and data
    header = rows[0]
    data = rows[2:]
    return header, data

def fappend(str,sheet):
    wb = load_workbook('output.xlsx')

    sheet = wb[sheet]  
    last_row_index = sheet.max_row


    column_index = 1  
    row_index = last_row_index + 1  

    cell = sheet.cell(row=row_index, column=column_index)
    cell.value = str 


    wb.save('output.xlsx')
    return 


def fappend_csv(rmd_table_string,to_return=False,sheet=''):

    if(rmd_table_string.find('|')):
        rmd_table_string=rmd_table_string[rmd_table_string.find('|'):]

    header, data = parse_rmd_table(rmd_table_string)

    max_columns = max(len(header), max(len(row) for row in data))

    header = header+ [' '] * (max_columns - len(header))

    data = [row + [''] * (max_columns - len(row)) for row in data]



    data_frame = pd.DataFrame(data, columns=header)
    wb = load_workbook('output.xlsx')

    sheet = wb[sheet]
    last_row_index = sheet.max_row
    rows = list(dataframe_to_rows(data_frame, index=False))
    for row in rows:
        sheet.append(row)

    wb.save('output.xlsx')

    


    return data_frame


def main_result(is_llama_parse=False):
    all_urls=get_all_companys_data()
    current_year=2024

    bofa=all_urls['BOFA']
    mgs=all_urls['MGS']

    jpm=all_urls['JPM']
    gs=all_urls['GS']

    #print(jpm)

    if(bofa):
        for year in range(current_year,current_year-3,-1):
            if(year in bofa):
                for quarter in bofa[year]:
                    if(quarter['isAvailabe']):
                        print(f"Year: {year}   Quarter: {quarter['Quarter']}")
                        fappend(f"\nYear: {year} , Quarter: {quarter['Quarter']}\n",'BOFA')
                        process(quarter['FinancialSuppliments'],bofa_search_words_with_threshold,bofa_prompt,is_llama_parse,year,quarter['Quarter'],'BOFA')
    if(mgs):

        for year in range(current_year,current_year-3,-1):
            if(year in mgs):
                for quarter in mgs[year]:
                    if(quarter['isAvailabe']):
                        print(f"Year: {year}   Quarter: {quarter['Quarter']}",quarter,'MGS')
                        fappend(f"\nYear: {year},   Quarter: {quarter['Quarter']}\n",'MGS')
                        process(quarter['FinancialSuppliments'],mgs_search_words_with_threshold,mgs_prompt,is_llama_parse,year,quarter['Quarter'],'MGS')
    
    if(jpm):
         #print(jpm)
          
         for year in range(current_year,current_year-3,-1):
            if(year in jpm):
                for quarter in jpm[year]:
                    if(quarter['isAvailabe']):
                        print(f"Year: {year}   Quarter: {quarter['Quarter']}",quarter)
                        fappend(f"\nYear: {year},   Quarter: {quarter['Quarter']}\n",'JPM')
                        process(quarter['FinancialSuppliments'],jpm_search_words_with_threshold,jpm_prompt,is_llama_parse,year,quarter['Quarter'],'JPM')


    if(gs):
        for year in range(current_year,current_year-3,-1):
             if(year in gs):
                for quarter in gs[year]:
                    if(quarter['isAvailabe']):
                        print(f"Year: {year}   Quarter: {quarter['Quarter']}",quarter)
                        fappend(f"\nYear: {year},   Quarter: {quarter['Quarter']}\n",'GS')
                        process(quarter['FinancialSuppliments'],gs_search_words_with_threshold,gs_prompt,is_llama_parse,year,quarter['Quarter'],'GS')




def process(url,thresh,prompt,is_llama_parse,year='',quarter='',company=''):
    text=""
    matching_pages,text=extract_pages_with_text_pypdf(url,thresh)
    print(matching_pages,url)
    if(is_llama_parse):
        matching_pages,text=extract_pages_with_text_llama_parse(url,thresh,is_llama_parse)

    timeline=''

    if(year and quarter):
        timeline= f"Extract for the year {year} and quarter{quarter} / Q{quarter}"


    # if(not matching_pages):
    #     print("---------")
    #     print(url,thresh)
    #     print("---------")
    #     return 
    # else:
    #     print(matching_pages,url)
    

    return invoke(text,prompt,timeline,company)

    


def invoke(text,prompt,timeline='',comany=''):

    data=generate_data(text,prompt,timeline)
    body=generate_body(data)
    return invoke_claude_model(bedrock_client,body,comany)


def web_user_interface():

    prompt = st.text_area('prompt', gs_prompt)

    gs_url='https://www.goldmansachs.com/media-relations/press-releases/current/pdfs/2023-q2-results.pdf'


    _url= st.text_area('url',gs_url)

   

    strmk=f'''
    <object data="{_url}" type="application/pdf" width="700px" height="700px">
        <embed src="{_url}">
            <p>This browser does not support PDFs. Please download the PDF to view it: <a href="{_url}">Download PDF</a>.</p>
        </embed>
    </object>
    '''

    st.markdown(strmk,True)
    

    _threshold=ast.literal_eval(st.text_area('thresholds',gs_search_words_with_threshold))
    print(type(_threshold))
    locakfilemk=f'''
            <object id="bla" data="file:///Users/rohan/Documents/MSU/Classes/MTH844/Code/temp.pdf" type="application/pdf" width="700px" height="700px">
                <embed src="file:///Users/rohan/Documents/MSU/Classes/MTH844/Code/temp.pdf">
                    <p>This browser does not support PDFs. Please download the PDF to view it: <a href="file:///Users/rohan/Documents/MSU/Classes/MTH844/Code/temp.pdf">Download PDF</a>.</p>
                </embed>
            </object>
            '''

   
    st.markdown(locakfilemk,True)
    extract_pages_with_text_llama_parse(_url,_threshold)
    getrpages=st.button('getpages')

    if(getrpages):
        extract_pages_with_text_llama_parse(_url,_threshold)


    reader=st.radio('reader:', ['PyPDF2','Llama Parse'])
    final_text=''

    if(reader=='Llama Parse'):
        pages_to_save,text=extract_pages_with_text_llama_parse(_url,_threshold,True)
        final_text=text

        st.text_area('string extracted',text)
    else:
        pages_to_save,text=extract_pages_with_text_pypdf(_url,_threshold)
        final_text=text
        st.text_area('string extracted',text)

    getresults=st.button('Run Model')

    if(getresults):
        df=invoke(final_text,prompt)
        st.table(df)



    
if __name__ == "__main__":
    bedrock_client= get_runtime()

    main_result(is_llama_parse=False)
    quarter="https://www.jpmorganchase.com//content/dam/jpmc/jpmorgan-chase-and-co/investor-relations/documents/quarterly-earnings/2023/4th-quarter/16d9371e-30e9-4898-abf6-d1f7c86fd311.pdf"

    #process(quarter,jpm_search_words_with_threshold,jpm_prompt,False,2023,4)


    #print(get_all_companys_data())

    #web_user_interface()

    #print(get_all_companys_data())
